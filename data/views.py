from django.shortcuts import render
from . models import EmpModel, JobModel, JobStarting
from django.template.loader import render_to_string
from .utils import render_to_pdf
from datetime import datetime
from django.http import HttpResponse
from openpyxl.styles import Alignment  , Font, Fill
import openpyxl
def index(request):
    emp=EmpModel.objects.count()
    jobs=JobModel.objects.count()
    context = {
        'emp':emp,
        'job':jobs
        }
    return render(request, "index.html", context)
def employee(request):
    if request.method=='POST':
        emp_id=request.POST['emp_id']
        emp_fname=request.POST['emp_fname']
        emp_lname=request.POST['emp_lname']
        emp_phone=request.POST['emp_phone']
        emp_email=request.POST['emp_email']
        emp=EmpModel.objects.create(emp_id=emp_id,first_name=emp_fname,last_name=emp_lname,phone=emp_phone, email=emp_email)
        print("Data Added")
    return render(request, "employee.html")
def jobs(request):
    if request.method=='POST':
        job_id=request.POST['job_id']
        job_title=request.POST['job_title']
        job_desc=request.POST['job_desc']
        job_sdate=request.POST['job_sdate']
        job_edate=request.POST['job_edate']
        job_status=request.POST['job_status']
        job=JobModel.objects.create(job_id=job_id, job_title=job_title, job_desc=job_desc, job_starting=job_sdate, job_ending=job_edate, job_status=job_status)
        print("Success")
    return render(request, "jobs.html")
def startingtime(request):
    employee=EmpModel.objects.all()
    jobs=JobModel.objects.all()
    if request.method=='POST':
        s_date=request.POST['s_date']
        s_name=request.POST['s_name']
        s_job=request.POST['s_job']
        s_start=request.POST['s_start']
        s_end=request.POST['s_end']
        s_lunch=request.POST['s_lunch']
        s_hollyday=request.POST['s_hollyday']
        s_status=request.POST['s_status']
        starting=JobStarting.objects.create(s_date=s_date, s_name=s_name, s_jobs=s_job, s_start=s_start, s_end=s_end, s_status=s_status, s_lunch=s_lunch,s_holyday=s_hollyday)
        print("Created Success")
    return render(request, "startingtime.html",{'employee':employee,'jobs':jobs})

def ResultList(request):
    th=0
    ot=0
    if request.method=='POST':
        if 'pdf' in request.POST:
            s_start=request.POST['s_start']
            s_end=request.POST['s_end'] 
            data=JobStarting.objects.filter(s_date__range=[s_start, s_end])
            for a in data:
                s1 = str(a.s_start)
                s2 = str(a.s_end)
                start_dt = datetime.strptime(s1, '%H:%M:%S')
                end_dt = datetime.strptime(s2, '%H:%M:%S')
                diff = (end_dt - start_dt)
                days = diff.days
                days_to_hours = days * 24
                diff_btw_two_times = (diff.seconds) / 3600
                overall_hours = days_to_hours + diff_btw_two_times
                if a.s_lunch=='Yes':
                    overall_hours=overall_hours-1
                th=th+overall_hours
            for a in data:
                s1 = str(a.s_start)
                s2 = str(a.s_end)
                start_dt = datetime.strptime(s1, '%H:%M:%S')
                end_dt = datetime.strptime(s2, '%H:%M:%S')
                diff = (end_dt - start_dt)
                days = diff.days
                days_to_hours = days * 24
                diff_btw_two_times = (diff.seconds) / 3600
                overall_hours = days_to_hours + diff_btw_two_times
                if a.s_lunch=="Yes" and a.s_holyday=="No" and overall_hours>8:
                    otdata=overall_hours-8
                    ot=ot+otdata
                if a.s_lunch=="No" and a.s_holyday=="No" and overall_hours>8:
                    otdata=overall_hours-8
                    ot=ot+otdata+1
                if a.s_lunch=="Yes" and a.s_holyday=="Yes":
                    otdata=overall_hours-1
                    ot=ot+otdata
                if a.s_lunch=="No" and a.s_holyday=="Yes":
                    otdata=overall_hours
                    ot=ot+otdata
            template_name = "pdf.html"
            return render_to_pdf(
            template_name,
            {
                "record": data,
                "date1":s_start,
                "date2":s_end,
                "overtime":ot,
                "workhour":th
            },
            )
        if 'excel' in request.POST:
            s_start=request.POST['s_start']
            s_end=request.POST['s_end'] 
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="mydata.xlsx"'

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'Report'
            worksheet.merge_cells('A1:G1')  
            cell = worksheet.cell(row=1, column=1)  
            cell.value = 'Company Name'  
            cell.alignment = Alignment(horizontal='center', vertical='center') 
            cell.font=Font(size=20, bold=True)
            worksheet.merge_cells('A2:C2')  
            cell = worksheet.cell(row=2, column=1)  
            cell.value = 'Starting Date:'+str(s_start)
            cell.alignment = Alignment(horizontal='center', vertical='center') 
            cell.font=Font(size=12, bold=True) 
            worksheet.merge_cells('D2:G2')  
            cell = worksheet.cell(row=2, column=4)  
            cell.value = 'Ending Date:'+str(s_end)
            cell.alignment = Alignment(horizontal='center', vertical='center') 
            cell.font=Font(size=12, bold=True) 
            worksheet.row_dimensions[1].height = 70
            worksheet.row_dimensions[2].height = 30
            worksheet.column_dimensions['A'].width = 20
            worksheet.column_dimensions['B'].width = 20
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 15
            worksheet.column_dimensions['E'].width = 15
            worksheet.column_dimensions['F'].width = 15
            worksheet.column_dimensions['G'].width = 15
            # Write header row
            header = ['Name', 'Job', 'Date', 'Starting Time', 'Ending Time', 'Working Hour','Over Time']
            for col_num, column_title in enumerate(header, 1):
                cell = worksheet.cell(row=3, column=col_num)
                cell.value = column_title
                cell.font=Font(bold=True)

            # Write data rows
            queryset = JobStarting.objects.filter(s_date__range=[s_start, s_end]).values_list('s_name', 's_jobs', 's_date', 's_start','s_end')
            for row_num, row in enumerate(queryset, 3):
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num+1, column=col_num)
                    cell.value = cell_value
            
            workbook.save(response)
            return response
        if 'table' in request.POST:
            s_start=request.POST['s_start']
            s_end=request.POST['s_end'] 
            data=JobStarting.objects.filter(s_date__range=[s_start, s_end])
            return render(request, "table.html",{"job":data})
    return render(request, "table.html")